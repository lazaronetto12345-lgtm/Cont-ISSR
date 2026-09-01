# ============================================================
#  APP ISSR — v28.7.2 (Cloud-Ready / Login Clean Branco)
#  LOGIN: ifesbiomol / biomol102030
#  Estrutura: Desktop/ssr_resultados/Cultura/Primer/arquivos
#  Nova pasta: Codificacao de bandas (Excel + TXT Ent-<Cultura>.txt)
#  EXCEL EM BLOCOS MULTIPRIMER: Arial 11, Centralizado
#  MODO CRIACAO LIVRE DE COLUNAS (clicar, arrastar, deletar)
#  Grava JPG mesmo em caminhos com acento (OneDrive/Area de Trabalho)
#  Filtros: Padrao, Lilas/Roxo, P&B, P&B Invertido, Amarelo Ouro
# ============================================================

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.cluster.hierarchy import linkage, dendrogram
from scipy.spatial.distance import squareform
from scipy.ndimage import uniform_filter1d
import cv2
import io
import base64
import os
import tempfile
import hashlib
import pickle
import openpyxl
from datetime import datetime

# ============================================================
#  CREDENCIAIS PADRAO DE ACESSO
# ============================================================
USUARIO_PADRAO = "ifesbiomol"
SENHA_PADRAO = "biomol102030"

st.set_page_config(page_title="ISSR Pro v28.7.2", page_icon=None, layout="wide")

# ============================================================
#  CSS DE AJUSTE DE LAYOUT, DROPDOWN E COR AZUL GLOBAL
# ============================================================
st.markdown("""
<style>
    /* ========================================================== */
    /* FORÇAR A COR AZUL (#007bff) NO STREAMLIT INTEIRO         */
    /* ========================================================== */
    
    /* 1. BOTÕES PRIMÁRIOS (ex: "Salvar em Codificacao de bandas") */
    button[kind="primary"] {
        background-color: #007bff !important;
        border-color: #007bff !important;
        color: #ffffff !important;
    }
    button[kind="primary"]:hover {
        background-color: #0056b3 !important;
        border-color: #0056b3 !important;
        color: #ffffff !important;
    }

    /* 2. TAGS DO MULTISELECT (As caixinhas de seleção vermelhas) */
    span[data-baseweb="tag"] {
        background-color: #007bff !important;
        color: #ffffff !important;
    }
    span[data-baseweb="tag"] span {
        color: #ffffff !important;
    }
    span[data-baseweb="tag"] svg {
        fill: #ffffff !important;
    }
    /* Efeito hover no 'X' para fechar a tag */
    span[data-baseweb="tag"] span[role="button"]:hover {
        background-color: #0056b3 !important;
    }

    /* 3. ABAS SUPERIORES (Tabs - Linha e Texto) */
    div[data-baseweb="tab-highlight"] {
        background-color: #007bff !important;
    }
    button[role="tab"][aria-selected="true"] {
        color: #007bff !important;
    }
    button[role="tab"][aria-selected="true"] div,
    button[role="tab"][aria-selected="true"] span,
    button[role="tab"][aria-selected="true"] p {
        color: #007bff !important;
    }

    /* 4. CHECKBOXES (Quadradinhos de marcar) */
    div[data-testid="stCheckbox"] label div[data-checked="true"] {
        background-color: #007bff !important;
        border-color: #007bff !important;
    }
    
    /* 5. SLIDERS (O controle de zoom px) */
    div[data-testid="stSlider"] div[role="slider"] {
        background-color: #007bff !important;
    }
    div[data-testid="stSlider"] div[data-baseweb="slider"] > div > div {
        background-color: #007bff !important;
    }

    /* ========================================================== */
    /* AJUSTES DE LAYOUT E DROPDOWNS ORIGINAIS DO PROJETO       */
    /* ========================================================== */
    .block-container { padding-top:1rem; padding-bottom:1rem; max-width:100% !important; }
    iframe { border:none !important; }
    div[data-testid="column"] button { width:100%; margin-bottom:4px; }
    
    div[data-testid="stMultiSelect"] div[data-baseweb="select"] > div {
        max-height: 95px !important;
        overflow-y: auto !important;
        flex-wrap: wrap !important;
        align-content: flex-start !important;
    }
    
    div[data-testid="stMultiSelect"] {
        margin-bottom: 10px !important;
    }

    div[data-baseweb="popover"],
    div[data-baseweb="layer"] {
        z-index: 999999 !important;
    }
    
    div[data-baseweb="popover"] ul[role="listbox"],
    div[data-baseweb="popover"] [role="listbox"] {
        max-height: 220px !important;
        overflow-y: auto !important;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
#  CARREGADOR DE LOGO VIA BASE64 (Garantido em qualquer pasta)
# ============================================================

def _carregar_logo_base64():
    """Localiza a imagem do logo na pasta do app.py e converte em Base64."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    extensoes = [".jpeg", ".jpg", ".png", ".webp"]
    
    # 1. Tenta encontrar por nome direto
    for ext in extensoes:
        caminho = os.path.join(base_dir, f"logo{ext}")
        if os.path.isfile(caminho):
            try:
                with open(caminho, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                    mime = "jpeg" if ext in [".jpg", ".jpeg"] else ext.replace(".", "")
                    return f"data:image/{mime};base64,{b64}"
            except Exception:
                pass

    # 2. Busca qualquer imagem na pasta que contenha 'logo' ou 'whatsapp'
    try:
        for f in os.listdir(base_dir):
            nome_lower = f.lower()
            if any(nome_lower.endswith(ext) for ext in extensoes):
                if "logo" in nome_lower or "whatsapp" in nome_lower:
                    caminho = os.path.join(base_dir, f)
                    with open(caminho, "rb") as img_file:
                        b64 = base64.b64encode(img_file.read()).decode()
                        ext_found = os.path.splitext(f)[1].lower()
                        mime = "jpeg" if ext_found in [".jpg", ".jpeg"] else ext_found.replace(".", "")
                        return f"data:image/{mime};base64,{b64}"
    except Exception:
        pass

    return None


# ============================================================
#  SISTEMA DE VERIFICACAO DE USUARIO E SENHA (LOGIN CLEAN)
# ============================================================

def verificar_autenticacao():
    if "autenticado" not in st.session_state:
        st.session_state["autenticado"] = False

    if st.session_state["autenticado"]:
        return True

    st.markdown("""
    <style>
        /* Esconde header do Streamlit */
        [data-testid="stHeader"] { display: none !important; }

        /* Fundo branco */
        [data-testid="stAppViewContainer"], .stApp, [data-testid="stAppViewContainer"] > .main {
            background: #ffffff !important;
            background-color: #ffffff !important;
        }

        /* Centraliza vertical e horizontal */
        [data-testid="stAppViewContainer"] > .main {
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            min-height: 100vh !important;
        }

        /* 
           Largura ESTREITA da caixa de login
           (precisa vencer o max-width:100% do CSS global do app)
        */
        [data-testid="block-container"],
        .block-container,
        .main .block-container {
            max-width: 360px !important;
            width: 360px !important;
            margin-left: auto !important;
            margin-right: auto !important;
            padding-top: 0 !important;
            padding-bottom: 0 !important;
            padding-left: 0 !important;
            padding-right: 0 !important;
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;

            /* Zoom 110% */
            transform: scale(1.1) !important;
            transform-origin: center center !important;
        }

        /* Inputs e botão ocupam só a largura da caixa (360px) */
        div[data-testid="stTextInput"],
        div[data-testid="stTextInput"] > div,
        div[data-testid="stTextInput"] input,
        div[data-testid="stButton"],
        div[data-testid="stButton"] button {
            width: 100% !important;
            max-width: 360px !important;
        }

        .titulo-sigpesq {
            color: #007bff;
            font-weight: 900;
            font-size: 26px;
            text-align: center;
            margin: 14px 0 28px 0;
            font-family: Arial, sans-serif;
            letter-spacing: 1px;
        }

        div[data-testid="stTextInput"] {
            margin-bottom: 8px !important;
        }

        button[kind="primary"] {
            background-color: #007bff !important;
            border-color: #007bff !important;
            border-radius: 6px !important;
            font-weight: bold !important;
            padding: 0.65rem !important;
            font-size: 16px !important;
            margin-top: 10px !important;
        }
        button[kind="primary"]:hover {
            background-color: #0056b3 !important;
            border-color: #0056b3 !important;
        }
    </style>
    """, unsafe_allow_html=True)

    # Sem colunas — evita esticar
    logo_b64 = _carregar_logo_base64()
    if logo_b64:
        st.markdown(
            f'<div style="text-align:center;">'
            f'<img src="{logo_b64}" style="max-height:130px;width:auto;border-radius:6px;"/>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.markdown('<div class="titulo-sigpesq">ISSR</div>', unsafe_allow_html=True)

    usuario_digitado = st.text_input(
        "Usuario", placeholder="Usuário", label_visibility="collapsed", key="login_user"
    )
    senha_digitada = st.text_input(
        "Senha", type="password", placeholder="Senha", label_visibility="collapsed", key="login_pass"
    )

    if st.button("Entrar", type="primary", use_container_width=True):
        if usuario_digitado == USUARIO_PADRAO and senha_digitada == SENHA_PADRAO:
            st.session_state["autenticado"] = True
            st.rerun()
        else:
            st.error("Credenciais inválidas!")

    return False

if not verificar_autenticacao():
    st.stop()


# ============================================================
#  SISTEMA DE CAMINHOS — Suporta OneDrive, acentos e Nuvem
# ============================================================
BASE_DIR = ""
CULTURA_DIR = ""
BACKUP_FILE = ""

def _is_cloud():
    return (
        os.environ.get("STREAMLIT_SERVER_HEADLESS") == "true"
        or os.path.exists("/home/appuser")
        or os.environ.get("HOSTNAME", "").startswith("streamlit")
    )

def abrir_pasta(path):
    if _is_cloud():
        return
    try:
        os.startfile(path)
    except Exception:
        try:
            import subprocess
            subprocess.Popen(["xdg-open", path])
        except Exception:
            pass

def _nome_arquivo_seguro(texto):
    t = str(texto).strip()
    return "".join(c if (c.isalnum() or c in " _-") else "_" for c in t).strip().replace(" ", "_") or "padrao"

def _get_desktop_path():
    home = os.path.expanduser("~")
    try:
        import ctypes
        from ctypes import wintypes, windll
        buf = ctypes.create_unicode_buffer(wintypes.MAX_PATH)
        if windll.shell32.SHGetFolderPathW(None, 0x0000, None, 0, buf) == 0:
            if buf.value and os.path.isdir(buf.value):
                return buf.value
    except Exception:
        pass

    userprofile = os.environ.get("USERPROFILE") or home
    for p in [
        os.path.join(userprofile, "OneDrive", "Area de Trabalho"),
        os.path.join(userprofile, "OneDrive", "Desktop"),
        os.path.join(userprofile, "OneDrive - Pessoal", "Area de Trabalho"),
        os.path.join(userprofile, "Area de Trabalho"),
        os.path.join(userprofile, "Desktop"),
        os.path.join(home, "Desktop"),
    ]:
        if os.path.isdir(p):
            return p

    fallback = os.path.join(userprofile, "Desktop")
    os.makedirs(fallback, exist_ok=True)
    return fallback

def atualizar_caminhos(nome_cultura):
    global BASE_DIR, CULTURA_DIR, BACKUP_FILE
    desktop_path = _get_desktop_path()
    nome_seguro = _nome_arquivo_seguro(nome_cultura) if nome_cultura else "Projeto_Sem_Nome"
    BASE_DIR = os.path.join(desktop_path, "ssr_resultados")
    CULTURA_DIR = os.path.join(BASE_DIR, nome_seguro)
    BACKUP_FILE = os.path.join(CULTURA_DIR, f"backup_{nome_seguro}.pkl")

def garantir_pastas_base():
    if not BASE_DIR or not CULTURA_DIR:
        atualizar_caminhos(st.session_state.get("nome_cultura_salva", "Minha_Cultura"))
    os.makedirs(BASE_DIR, exist_ok=True)
    os.makedirs(CULTURA_DIR, exist_ok=True)
    return os.path.abspath(CULTURA_DIR)

def garantir_pasta_primer(nome_primer):
    garantir_pastas_base()
    pasta_primer = os.path.join(CULTURA_DIR, _nome_arquivo_seguro(nome_primer))
    os.makedirs(pasta_primer, exist_ok=True)
    return os.path.abspath(pasta_primer)

def _salvar_jpg_unicode(caminho, img_bgr, qualidade=95):
    try:
        ok, buf = cv2.imencode(".jpg", img_bgr, [int(cv2.IMWRITE_JPEG_QUALITY), int(qualidade)])
        if not ok:
            return False, "imencode retornou False"
        with open(caminho, "wb") as f:
            f.write(buf.tobytes())
        if not (os.path.isfile(caminho) and os.path.getsize(caminho) > 0):
            return False, "arquivo nao encontrado apos gravacao"
        return True, "ok"
    except Exception as e:
        return False, str(e)

def salvar_gel_conferencia(img_bgr, nome_primer, bandas_salvas=None, lista_f1=None, lista_f2=None,
                            edges1=None, edges2=None, w1=0, skip1=0, skip2=0, y_guia=0):
    try:
        pasta_primer = garantir_pasta_primer(nome_primer)
        if img_bgr is None:
            return False, "Imagem base do gel nao encontrada."

        img = np.ascontiguousarray(img_bgr.copy())
        h, w = img.shape[:2]
        lista_f1 = lista_f1 or []
        lista_f2 = lista_f2 or []
        bandas_salvas = bandas_salvas or []
        edges1 = edges1 or []
        edges2 = edges2 or []

        if y_guia and 0 < int(y_guia) < h:
            cv2.line(img, (0, int(y_guia)), (w - 1, int(y_guia)), (0, 255, 120), 2)

        for i, b in enumerate(bandas_salvas):
            y = int(round(float(b.get("y", -1))))
            if 0 <= y < h:
                cv2.line(img, (0, y), (w - 1, y), (200, 200, 200), 1)
                letra = chr(97 + i) if i < 26 else f"b{i}"
                cv2.putText(img, letra, (10, max(20, y - 4)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)

        def _centro_lane(edges, x_off, lane):
            if not edges or lane < 0 or lane >= len(edges) - 1:
                return None, None
            x0 = float(edges[lane]) + x_off
            x1 = float(edges[lane + 1]) + x_off
            return (x0 + x1) / 2.0, max(6.0, (x1 - x0) * 0.75)

        for b in bandas_salvas:
            y = int(round(float(b.get("y", -1))))
            marks = b.get("marks") or []
            if y < 0 or y >= h:
                continue
            for col_idx, m in enumerate(marks):
                if int(m) != 1:
                    continue
                cx, tw = None, 20
                if col_idx < len(lista_f1):
                    lane = col_idx + int(skip1)
                    cx, tw = _centro_lane(edges1, 0, lane)
                else:
                    idx2 = col_idx - len(lista_f1)
                    if 0 <= idx2 < len(lista_f2):
                        lane = idx2 + int(skip2)
                        cx, tw = _centro_lane(edges2, float(w1), lane)
                if cx is None:
                    continue
                x1_r = int(cx - tw / 2)
                x2_r = int(cx + tw / 2)
                y1_r = max(0, y - 3)
                y2_r = min(h - 1, y + 3)
                cv2.rectangle(img, (x1_r, y1_r), (x2_r, y2_r), (0, 255, 0), -1)
                cv2.rectangle(img, (x1_r, y1_r), (x2_r, y2_r), (0, 80, 0), 1)

        safe = _nome_arquivo_seguro(nome_primer)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = os.path.join(pasta_primer, f"gel_{safe}_{ts}.jpg")

        ok, msg = _salvar_jpg_unicode(caminho, img, 95)
        if not ok:
            return False, f"Falha ao gravar JPG: {msg} — caminho: {caminho}"
        return True, os.path.abspath(caminho)
    except Exception as e:
        return False, str(e)

def salvar_matriz_primer(nome_primer, df_primer):
    try:
        pasta_primer = garantir_pasta_primer(nome_primer)
        safe = _nome_arquivo_seguro(nome_primer)
        caminho_xlsx = os.path.join(pasta_primer, f"Matriz_{safe}.xlsx")
        excel_data = exportar_excel_completo({nome_primer: df_primer}, nome_export=nome_primer)
        with open(caminho_xlsx, "wb") as f:
            f.write(excel_data)
        if not (os.path.isfile(caminho_xlsx) and os.path.getsize(caminho_xlsx) > 0):
            return False, f"Excel nao encontrado apos gravacao: {caminho_xlsx}"
        return True, os.path.abspath(caminho_xlsx)
    except Exception as e:
        return False, str(e)

def salvar_backup_global():
    try:
        garantir_pastas_base()
        with open(BACKUP_FILE, "wb") as f:
            pickle.dump(st.session_state["todas_matrizes"], f)
        return True, os.path.abspath(BACKUP_FILE)
    except Exception as e:
        return False, str(e)


# ============================================================
#  PASTA: CODIFICACAO DE BANDAS + TXT TRANSPOSTO
# ============================================================

def garantir_pasta_codificacao():
    garantir_pastas_base()
    pasta = os.path.join(CULTURA_DIR, "Codificacao de bandas")
    os.makedirs(pasta, exist_ok=True)
    return os.path.abspath(pasta)


def gerar_txt_transposto(primers_dict):
    dfs = list(primers_dict.values())
    if not dfs:
        return ""

    df_combined = pd.concat(dfs, axis=0)

    acessos = [str(c) for c in df_combined.columns.tolist()]
    if st.session_state.get("ordenar_ids_auto", True):
        acessos = ordenar_ids(acessos)
    df_combined = df_combined.reindex(columns=acessos).fillna(0)

    def _to01(x):
        try:
            if x is True or x == 1 or x == 1.0 or str(x).strip() == "1":
                return 1
            return 0
        except Exception:
            return 0

    try:
        df_bin = df_combined.map(_to01)
    except Exception:
        df_bin = df_combined.applymap(_to01)

    mat = df_bin.T.to_numpy()

    linhas = []
    n_bandas_ref = None
    for i in range(mat.shape[0]):
        tokens = []
        for j in range(mat.shape[1]):
            tokens.append("1" if int(mat[i, j]) == 1 else "0")

        linha = " ".join(tokens)
        linha = " ".join(linha.split())

        partes = linha.split(" ")
        if n_bandas_ref is None:
            n_bandas_ref = len(partes)
        if len(partes) != n_bandas_ref:
            raise ValueError(
                f"Linha {i+1} tem {len(partes)} bandas; esperado {n_bandas_ref}."
            )
        for p in partes:
            if p not in ("0", "1"):
                raise ValueError(f"Valor invalido '{p}' na linha {i+1}.")

        linhas.append(linha)

    return "\r\n".join(linhas) + "\r\n"


def salvar_codificacao_bandas(primers_dict, nome_cultura, dist_jaccard=None, acessos=None):
    try:
        pasta = garantir_pasta_codificacao()
        nome_seguro = _nome_arquivo_seguro(nome_cultura)

        nome_export = "Combinado" if len(primers_dict) > 1 else list(primers_dict.keys())[0]
        excel_data = exportar_excel_completo(primers_dict, nome_export, dist_jaccard, acessos)
        caminho_xlsx = os.path.join(pasta, f"SSR_{_nome_arquivo_seguro(str(nome_export))}.xlsx")
        with open(caminho_xlsx, "wb") as f:
            f.write(excel_data)

        txt_content = gerar_txt_transposto(primers_dict)
        caminho_txt = os.path.join(pasta, f"Ent-{nome_seguro}.txt")
        with open(caminho_txt, "wb") as f:
            f.write(txt_content.encode("ascii", errors="strict"))

        return True, pasta, os.path.abspath(caminho_xlsx), os.path.abspath(caminho_txt)
    except Exception as e:
        return False, str(e), None, None


def build_df_from_bands(bands, acessos, primer_name):
    row_names = []
    data_rows = []
    for b in bands:
        b_label = b["band"]
        if b_label.startswith(f"{primer_name}_"):
            idx_name = b_label
        else:
            idx_name = f"{primer_name}_{b_label}"
        row_names.append(idx_name)
        data_rows.append(b["vals"])
    df = pd.DataFrame(data_rows, index=row_names, columns=acessos)
    return df

def importar_excel_completo(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet_name = None
        for name in wb.sheetnames:
            if name.startswith("Matriz_"):
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]
        first_row = [cell.value for cell in ws[1]]
        if not first_row or len(first_row) < 2:
            return None, "A planilha parece estar vazia ou fora do formato padrao."

        first_primer = str(first_row[0]).strip()
        acessos = [str(val).strip() for val in first_row[1:] if val is not None]
        primers_dict = {}
        current_primer = first_primer
        current_bands = []

        for r_idx in range(2, ws.max_row + 1):
            row_cells = [ws.cell(row=r_idx, column=c_idx) for c_idx in range(1, len(acessos) + 2)]
            col1_val = row_cells[0].value
            vals = [c.value for c in row_cells[1:]]

            if col1_val is None and all(v is None for v in vals):
                continue
            if col1_val is not None and all(v is None for v in vals):
                if current_bands:
                    df = build_df_from_bands(current_bands, acessos, current_primer)
                    primers_dict[current_primer] = df
                current_primer = str(col1_val).strip()
                current_bands = []
                continue

            if col1_val is not None:
                band_name = str(col1_val).strip()
                binary_vals = []
                for v in vals:
                    if v is None:
                        binary_vals.append(False)
                    else:
                        try:
                            binary_vals.append(bool(int(float(v))))
                        except:
                            binary_vals.append(False)

                if len(binary_vals) < len(acessos):
                    binary_vals += [False] * (len(acessos) - len(binary_vals))
                else:
                    binary_vals = binary_vals[:len(acessos)]
                current_bands.append({"band": band_name, "vals": binary_vals})

        if current_bands:
            df = build_df_from_bands(current_bands, acessos, current_primer)
            primers_dict[current_primer] = df

        return primers_dict, None
    except Exception as e:
        return None, str(e)


# ============================================================
#  FUNCOES PYTHON DE ORDENACAO E PROCESSAMENTO
# ============================================================

def ordenar_ids(lista):
    if not lista:
        return []
    def chave_ordenacao(item):
        val = str(item).strip().upper()
        if val in ["L", "M", "LADDER", "MARCADOR"]:
            return (-1, 0, val)
        try:
            return (0, int(val), "")
        except ValueError:
            return (1, 0, val)
    return sorted(list(dict.fromkeys(lista)), key=chave_ordenacao)

def aplicar_filtro_bw(img_bgr, modo_filtro, brilho, contraste):
    if modo_filtro == "Padrao (Cor Original)":
        proc = img_bgr.copy()

    elif modo_filtro == "Lilas/Roxo (Fundo Roxo / Bandas Brancas)":
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        gray = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
        f = gray.astype(np.float32) / 255.0
        b = (150 + f * (255 - 150)).astype(np.uint8)
        g = (30  + f * (255 - 30)).astype(np.uint8)
        r = (90  + f * (255 - 90)).astype(np.uint8)
        proc = cv2.merge([b, g, r])

    elif modo_filtro == "Preto e Branco (Invertido - Fundo Branco)":
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        proc = cv2.cvtColor(cv2.bitwise_not(gray), cv2.COLOR_GRAY2BGR)

    elif modo_filtro == "Preto e Branco (Fundo Preto)":
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        proc = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)

    elif modo_filtro == "Amarelo (Fundo Amarelo / Bandas Pretas)":
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        gray = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)

        inv = cv2.bitwise_not(gray)
        f = inv.astype(np.float32) / 255.0

        f = np.power(f, 1.2)

        b = (0   * f + 10 * (1.0 - f)).astype(np.uint8)
        g = (160 * f + 20 * (1.0 - f)).astype(np.uint8)
        r = (255 * f + 40 * (1.0 - f)).astype(np.uint8)
        proc = cv2.merge([b, g, r])

    else:
        proc = img_bgr.copy()

    return cv2.convertScaleAbs(proc, alpha=contraste, beta=int((brilho - 1.0) * 50))


def transformar_imagem(img, angulo, escala, offset_y, cor_fundo):
    h, w = img.shape[:2]
    M = cv2.getRotationMatrix2D((w // 2, h // 2), angulo, escala)
    M[1, 2] += offset_y
    return cv2.warpAffine(img, M, (w, h), borderValue=cor_fundo)

def detectar_fundo_pocos(img_bgr):
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    y_max_busca = max(30, int(h * 0.28))
    roi = gray[:y_max_busca, :]
    roi_eq = cv2.equalizeHist(roi)
    roi_blur = cv2.GaussianBlur(roi_eq, (5, 5), 0)
    sobel = cv2.Sobel(roi_blur, cv2.CV_64F, 0, 1, ksize=3)
    sobel = np.abs(sobel)
    perfil = sobel.sum(axis=1)
    perfil = uniform_filter1d(perfil.astype(float), size=7)
    y0 = max(5, int(h * 0.03))
    y1 = y_max_busca - 2
    if y1 <= y0 + 5:
        return int(h * 0.10)
    trecho = perfil[y0:y1]
    limiar = np.percentile(trecho, 75)
    candidatos = []
    for i in range(2, len(trecho) - 2):
        if trecho[i] >= limiar and trecho[i] >= trecho[i - 1] and trecho[i] >= trecho[i + 1]:
            candidatos.append((trecho[i], y0 + i))
    if not candidatos:
        return int(y0 + np.argmax(trecho))
    candidatos.sort(reverse=True, key=lambda t: t[0])
    top = candidatos[:min(5, len(candidatos))]
    y_fundo = max(top, key=lambda t: t[1])[1]
    return int(min(h - 1, y_fundo + 2))

def auto_calibrar(img1_bgr, img2_bgr, rot1=0.0, rot2=0.0, escala2=1.0):
    cor = (0, 0, 0)
    i1 = transformar_imagem(img1_bgr, rot1, 1.0, 0, cor)
    i2 = transformar_imagem(img2_bgr, rot2, escala2, 0, cor)
    h1, w1_img = i1.shape[:2]
    h2, w2_img = i2.shape[:2]
    if h2 != h1:
        nw2 = int(w2_img * (h1 / h2))
        i2 = cv2.resize(i2, (nw2, h1), interpolation=cv2.INTER_AREA)
    y1 = detectar_fundo_pocos(i1)
    y2 = detectar_fundo_pocos(i2)
    return int(y1 - y2), int(y1), y1, y2

def calcular_jaccard(matriz):
    n = matriz.shape[0]
    dist = np.zeros((n, n))
    for i in range(n):
        for j in range(i + 1, n):
            a = int(np.sum((matriz[i] == 1) & (matriz[j] == 1)))
            b = int(np.sum((matriz[i] == 1) & (matriz[j] == 0)))
            c = int(np.sum((matriz[i] == 0) & (matriz[j] == 1)))
            d = a + b + c
            dist[i, j] = 0.0 if d == 0 else 1.0 - (a / d)
            dist[j, i] = dist[i, j]
    return dist

def fazer_upgma(dist_matrix):
    dm = dist_matrix.copy()
    np.fill_diagonal(dm, 0)
    dm = (dm + dm.T) / 2
    return linkage(squareform(dm, checks=False), method="average")

def plotar_dendrograma(Z, labels, titulo):
    fig, ax = plt.subplots(figsize=(max(10, len(labels) * 0.35), 7))
    dendrogram(Z, labels=labels, ax=ax, leaf_rotation=90, leaf_font_size=8,
               color_threshold=0.7 * max(Z[:, 2]) if len(Z) > 0 else 0)
    ax.set_title(titulo, fontsize=13, fontweight="bold")
    ax.set_ylabel("Dissimilaridade de Jaccard")
    plt.tight_layout()
    return fig

def carregar_imagem(uploaded):
    try:
        data = uploaded.read()
        img = cv2.imdecode(np.frombuffer(data, np.uint8), cv2.IMREAD_COLOR)
        return img, None
    except Exception as e:
        return None, str(e)


# ============================================================
#  EXPORTACAO EXCEL EM BLOCOS (ARIAL 11, CENTRALIZADO)
# ============================================================

def exportar_excel_completo(primers_dict, nome_export="Combinado", dist_jaccard=None, acessos=None):
    from openpyxl.styles import Font, Alignment, PatternFill

    buf = io.BytesIO()

    if isinstance(primers_dict, pd.DataFrame):
        primers_list = [(nome_export, primers_dict)]
    elif isinstance(primers_dict, dict):
        primers_list = [(p_name, df) for p_name, df in primers_dict.items()]
    elif isinstance(primers_dict, list):
        primers_list = primers_dict
    else:
        primers_list = [(nome_export, primers_dict)]

    all_acessos = []
    for _, df_p in primers_list:
        for col in df_p.columns:
            if col not in all_acessos:
                all_acessos.append(col)

    if st.session_state.get("ordenar_ids_auto", True):
        all_acessos = ordenar_ids(all_acessos)

    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        sheet_matriz = f"Matriz_{nome_export}"[:31]
        df_init = pd.DataFrame(columns=all_acessos)
        df_init.to_excel(w, sheet_name=sheet_matriz, index=False)
        ws1 = w.sheets[sheet_matriz]
        ws1.delete_rows(1, ws1.max_row + 1)

        font_arial = Font(name="Arial", size=11, bold=False)
        font_arial_bold = Font(name="Arial", size=11, bold=True)
        alinhamento_centro = Alignment(horizontal="center", vertical="center")

        current_row = 1

        for idx_primer, (p_name, df_p) in enumerate(primers_list):
            if idx_primer == 0:
                cell_a1 = ws1.cell(row=current_row, column=1, value=p_name)
                cell_a1.font = font_arial_bold
                cell_a1.alignment = alinhamento_centro
                for c_idx, acc in enumerate(all_acessos, start=2):
                    cell = ws1.cell(row=current_row, column=c_idx, value=acc)
                    cell.font = font_arial_bold
                    cell.alignment = alinhamento_centro
                current_row += 1
            else:
                cell_p = ws1.cell(row=current_row, column=1, value=p_name)
                cell_p.font = font_arial_bold
                cell_p.alignment = alinhamento_centro
                current_row += 1

            for band_name in df_p.index:
                s_band = str(band_name)
                band_label = s_band.split("_")[-1] if "_" in s_band else s_band

                cell_b = ws1.cell(row=current_row, column=1, value=band_label)
                cell_b.font = font_arial_bold
                cell_b.alignment = alinhamento_centro

                for c_idx, acc in enumerate(all_acessos, start=2):
                    val = 0
                    if acc in df_p.columns:
                        v = df_p.loc[band_name, acc]
                        val = 1 if bool(v) else 0
                    cell_v = ws1.cell(row=current_row, column=c_idx, value=val)
                    cell_v.font = font_arial
                    cell_v.alignment = alinhamento_centro
                current_row += 1
            current_row += 1

        for col in ws1.columns:
            ml = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws1.column_dimensions[col_letter].width = max(ml + 3, 6)

        if dist_jaccard is not None and acessos is not None:
            df_dist = pd.DataFrame(dist_jaccard, index=acessos, columns=acessos)
            df_dist.to_excel(w, sheet_name="Dissimilaridade_Jaccard")
            ws2 = w.sheets["Dissimilaridade_Jaccard"]
            ws2.cell(row=1, column=1, value="Matriz")

            for row in ws2.iter_rows():
                for cell in row:
                    cell.font = font_arial
                    cell.alignment = alinhamento_centro
                    cell.fill = PatternFill(fill_type=None)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.000'

            for cell in ws2[1]:
                cell.font = font_arial_bold
            for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=1):
                row[0].font = font_arial_bold

            for col in ws2.columns:
                ml = max(len(str(cell.value or "")) for cell in col)
                col_letter = col[0].column_letter
                ws2.column_dimensions[col_letter].width = max(ml + 3, 8)

    return buf.getvalue()


# ============================================================
#  HTML CANVAS - ATUALIZADO PARA NUVEM
# ============================================================

HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<script>
(function(){
    var S={
        setComponentReady:function(){ window.parent.postMessage({isStreamlitMessage:true,type:"streamlit:componentReady",apiVersion:1},"*"); },
        setFrameHeight:function(h){ window.parent.postMessage({isStreamlitMessage:true,type:"streamlit:setFrameHeight",height:h},"*"); },
        setComponentValue:function(v){ window.parent.postMessage({isStreamlitMessage:true,type:"streamlit:setComponentValue",value:v},"*"); },
        events:{addEventListener:function(type,cb){ window.addEventListener("message",function(ev){ if(ev.data&&ev.data.type===type) cb({detail:ev.data}); }); }}
    };
    window.Streamlit=S;
})();
</script>
<style>
*{box-sizing:border-box;}
body{margin:0;padding:0;background:#111;font-family:'Segoe UI',sans-serif;user-select:none;overflow:hidden;}
#calib-panel{background:linear-gradient(135deg,#1a252f,#2c3e50);border-bottom:2px solid #3498db;padding:10px 16px;display:flex;align-items:center;gap:12px;flex-wrap:wrap;font-size:12px;color:#ecf0f1;}
.btn-edit{padding:7px 16px;border:2px solid #3498db;border-radius:6px;background:#3498db;color:#fff;font-weight:bold;font-size:12px;cursor:pointer;white-space:nowrap;}
.btn-edit:hover{background:#2980b9;}
.btn-edit.ativo{background:#e74c3c;border-color:#e74c3c;animation:pulse 1.5s infinite;}
.btn-clear{padding:7px 16px;border:2px solid #e74c3c;border-radius:6px;background:#c0392b;color:#fff;font-weight:bold;font-size:12px;cursor:pointer;white-space:nowrap;}
.btn-clear:hover{background:#e74c3c;}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.8}}
#calib-instrucao{flex:1;padding:6px 12px;background:rgba(52,152,219,0.15);border:1px solid rgba(52,152,219,0.4);border-radius:5px;color:#3498db;font-weight:600;font-size:12px;min-height:30px;display:flex;align-items:center;}
#toolbar{padding:7px 14px;background:linear-gradient(135deg,#2c3e50,#34495e);color:#ecf0f1;border-bottom:2px solid #1a252f;font-size:11px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:5px;}
.tb-group{display:flex;align-items:center;gap:7px;flex-wrap:wrap;}
.ci{display:flex;align-items:center;gap:4px;padding:3px 8px;background:rgba(255,255,255,0.08);border-radius:4px;}
.lbl{font-weight:600;color:#3498db;font-size:11px;}
.val{color:#ecf0f1;font-size:11px;}
#container-wrapper{width:100%;height:520px;overflow:auto;background:#0d0d0d;}
#gel-canvas{cursor:crosshair;display:block;}
#container-wrapper::-webkit-scrollbar{width:10px;height:10px;}
#container-wrapper::-webkit-scrollbar-track{background:#2c3e50;}
#container-wrapper::-webkit-scrollbar-thumb{background:#3498db;border-radius:5px;}
#toast{position:fixed;bottom:14px;left:50%;transform:translateX(-50%);color:#fff;padding:8px 18px;border-radius:6px;font-size:12px;box-shadow:0 3px 10px rgba(0,0,0,0.4);z-index:9999;display:none;font-weight:600;pointer-events:none;}
</style>
</head>
<body>

<div id="calib-panel">
    <button id="btn-edit" class="btn-edit" onclick="toggleEditLanes()">Ajustar / Criar Colunas Manualmente</button>
    <button id="btn-clear" class="btn-clear" onclick="clearLanes()">Apagar Todas as Linhas</button>
    <div id="calib-instrucao">
        <strong>Zoom:</strong> roda do mouse · <strong>Mover:</strong> botao do meio OU Espaco+arrastar · <strong>Colunas:</strong> ative o modo azul
    </div>
</div>

<div id="toolbar">
    <div class="tb-group">
        <div class="ci"><span class="lbl">Roda:</span><span class="val">Zoom</span></div>
        <div class="ci"><span class="lbl">Meio/Espaco:</span><span class="val">Mover</span></div>
        <div class="ci"><span class="lbl">Esq.(vazio):</span><span class="val">Criar banda</span></div>
        <div class="ci"><span style="color:#0f0;font-weight:bold;">━</span><span class="lbl">Esq.(banda):</span><span class="val">Marcar</span></div>
        <div class="ci"><span class="lbl">Dir.:</span><span class="val">Excluir banda</span></div>
    </div>
    <div style="color:#2ecc71;font-weight:bold;font-size:11px;">Laser — Fundo dos Pocos</div>
</div>

<div id="toast"></div>
<div id="container-wrapper"><canvas id="gel-canvas"></canvas></div>

<script>
const S=window.Streamlit;
const container=document.getElementById('container-wrapper');
const canvas=document.getElementById('gel-canvas');
const ctx=canvas.getContext('2d');
const img=new Image();

let originalW=0, originalH=0, scale=1;
let yGuia=0, bandas=[], list1=[], list2=[], w1=0, w2=0, skip1=0, skip2=0;
let edges1=[], edges2=[], isReady=false;

let isEditingLanes=false;
let draggingEdge=null;
let hoverEdge=null;

let isPanning=false;
let panStartX=0, panStartY=0, scrollStartL=0, scrollStartT=0;
let spacePressed=false;

function toast(msg,type='success'){
    const t=document.getElementById('toast');
    const c={success:'rgba(46,204,113,.96)',info:'rgba(52,152,219,.96)',warning:'rgba(241,196,15,.96)',error:'rgba(231,76,60,.96)'};
    t.style.background=c[type]||c.success; t.textContent=msg; t.style.display='block';
    setTimeout(()=>{t.style.display='none';},2200);
}
function instrucao(html){ document.getElementById('calib-instrucao').innerHTML=html; }
function sendData(){ if(S) S.setComponentValue({bandas, calib_edges1:edges1, calib_edges2:edges2}); }

function clearLanes(){
    if(confirm('Deseja APAGAR todas as linhas e desenhar do zero?')){
        edges1=[]; edges2=[]; draw(); sendData();
        toast('Grade apagada!','info');
        if(!isEditingLanes) toggleEditLanes();
    }
}

function toggleEditLanes(){
    isEditingLanes=!isEditingLanes;
    const btn=document.getElementById('btn-edit');
    if(isEditingLanes){
        btn.className='btn-edit ativo';
        btn.textContent='Salvar Colunas e Voltar';
        instrucao('<strong>MODO COLUNAS:</strong> Esq=criar/arrastar linha · Dir=apagar linha · Meio/Espaco=mover · Roda=zoom');
        toast('Modo colunas ativo','info');
    } else {
        btn.className='btn-edit';
        btn.textContent='Ajustar / Criar Colunas Manualmente';
        instrucao('<strong>Zoom:</strong> roda · <strong>Mover:</strong> botao do meio OU Espaco+arrastar · <strong>Colunas:</strong> ative o modo azul');
        toast('Colunas salvas!','success');
        sendData();
    }
    draw();
}

function getNearestEdge(mX){
    let minDist=8/scale, nearest=null;
    if(edges1){
        for(let i=0;i<edges1.length;i++){
            const d=Math.abs(mX-edges1[i]);
            if(d<minDist){ minDist=d; nearest={group:1,idx:i}; }
        }
    }
    if(edges2){
        for(let i=0;i<edges2.length;i++){
            const d=Math.abs((mX-w1)-edges2[i]);
            if(d<minDist){ minDist=d; nearest={group:2,idx:i}; }
        }
    }
    return nearest;
}

function laneFromEdges(edges,x){
    if(!edges||edges.length<2) return -1;
    for(let k=0;k<edges.length-1;k++){
        if(x>=edges[k]&&x<edges[k+1]) return k;
    }
    if(x>=edges[edges.length-1]) return edges.length-2;
    return -1;
}

function draw(){
    if(!img.complete||originalW===0) return;
    canvas.width=Math.round(originalW*scale);
    canvas.height=Math.round(originalH*scale);
    ctx.imageSmoothingEnabled=true; ctx.imageSmoothingQuality='high';
    ctx.clearRect(0,0,canvas.width,canvas.height);
    ctx.drawImage(img,0,0,canvas.width,canvas.height);

    const xSplit=w1*scale;
    const e1=edges1?edges1.map(v=>v*scale):[];
    const e2=edges2?edges2.map(v=>xSplit+v*scale):[];

    function drawEdgeLine(xCanvas,isId,isHovered){
        ctx.save(); ctx.beginPath();
        ctx.moveTo(xCanvas,0); ctx.lineTo(xCanvas,canvas.height);
        if(isEditingLanes){
            ctx.strokeStyle=isHovered?'#e74c3c':'#3498db';
            ctx.lineWidth=isHovered?Math.max(2.5,3*scale):Math.max(1.5,1.5*scale);
            ctx.setLineDash([8,4]);
        } else {
            ctx.strokeStyle=isId?'rgba(255,220,50,0.7)':'rgba(255,255,255,0.15)';
            ctx.lineWidth=isId?Math.max(1,1.1*scale):Math.max(0.4,0.6*scale);
        }
        ctx.stroke(); ctx.restore();
    }

    for(let k=0;k<e1.length;k++){
        if(e1[k]<-1||e1[k]>xSplit+1) continue;
        const isId=(k>=skip1&&k<skip1+list1.length);
        const isHov=hoverEdge&&hoverEdge.group===1&&hoverEdge.idx===k;
        drawEdgeLine(e1[k],isId,isHov);
    }
    for(let k=0;k<e2.length;k++){
        if(e2[k]<xSplit-1||e2[k]>canvas.width+1) continue;
        const isId=(k>=skip2&&k<skip2+list2.length);
        const isHov=hoverEdge&&hoverEdge.group===2&&hoverEdge.idx===k;
        drawEdgeLine(e2[k],isId,isHov);
    }

    if(yGuia>0){
        const yL=yGuia*scale;
        ctx.save(); ctx.beginPath();
        ctx.moveTo(0,yL); ctx.lineTo(canvas.width,yL);
        ctx.strokeStyle='#2ecc71'; ctx.lineWidth=Math.max(1,1.5*scale);
        ctx.setLineDash([10,5]); ctx.stroke(); ctx.setLineDash([]); ctx.restore();
    }

    const yLaserPx=yGuia>0?yGuia*scale:canvas.height*0.15;
    const yId=Math.max(12,yLaserPx*0.55);

    function drawID(txt,cx,y,fs){
        ctx.save();
        ctx.font=`bold ${fs}px Arial`;
        ctx.textAlign='center'; ctx.textBaseline='middle';
        ctx.lineWidth=Math.max(2,2.5*scale);
        ctx.strokeStyle='rgba(0,0,0,0.95)'; ctx.strokeText(txt,cx,y);
        ctx.fillStyle='#ffffff'; ctx.fillText(txt,cx,y);
        ctx.restore();
    }

    if(e1.length>1){
        for(let i=0;i<list1.length;i++){
            const lane=i+skip1;
            if(lane<0||lane>=e1.length-1) continue;
            const cx=(e1[lane]+e1[lane+1])/2;
            const fs=Math.max(7,Math.min(13,Math.round((e1[lane+1]-e1[lane])*0.38)));
            if(cx<0||cx>xSplit) continue;
            drawID(list1[i],cx,yId,fs);
        }
    }
    if(e2.length>1){
        for(let i=0;i<list2.length;i++){
            const lane=i+skip2;
            if(lane<0||lane>=e2.length-1) continue;
            const cx=(e2[lane]+e2[lane+1])/2;
            const fs=Math.max(7,Math.min(13,Math.round((e2[lane+1]-e2[lane])*0.38)));
            if(cx<xSplit||cx>canvas.width) continue;
            drawID(list2[i],cx,yId,fs);
        }
    }

    bandas.forEach((b,index)=>{
        const yD=b.y*scale;
        const letra=index<26?String.fromCharCode(97+index):('b'+index);
        ctx.save(); ctx.beginPath(); ctx.moveTo(0,yD); ctx.lineTo(canvas.width,yD);
        ctx.strokeStyle='rgba(255,255,255,0.30)'; ctx.lineWidth=Math.max(0.8,1*scale);
        ctx.setLineDash([5,4]); ctx.stroke(); ctx.setLineDash([]); ctx.restore();

        const fB=Math.max(9,Math.min(14,Math.round(12*scale)));
        ctx.save(); ctx.font=`bold ${fB}px Arial`; ctx.textBaseline='bottom';
        ctx.lineWidth=3; ctx.strokeStyle='rgba(0,0,0,0.85)'; ctx.fillStyle='#ffff00';
        ctx.textAlign='left'; ctx.strokeText(letra,5,yD-2); ctx.fillText(letra,5,yD-2);
        ctx.textAlign='right'; ctx.strokeText(letra,canvas.width-5,yD-2); ctx.fillText(letra,canvas.width-5,yD-2);
        ctx.restore();

        if(b.marks){
            b.marks.forEach((mark,colIdx)=>{
                if(mark!==1) return;
                let cx=-1,tw=20*scale;
                if(colIdx<list1.length){
                    const lane=colIdx+skip1;
                    if(e1&&lane>=0&&lane<e1.length-1){
                        cx=(e1[lane]+e1[lane+1])/2;
                        if(cx<0||cx>xSplit) cx=-1;
                        else tw=Math.max(6,(e1[lane+1]-e1[lane])*0.75);
                    }
                } else {
                    const idx2=colIdx-list1.length, lane=idx2+skip2;
                    if(e2&&lane>=0&&lane<e2.length-1){
                        cx=(e2[lane]+e2[lane+1])/2;
                        if(cx<xSplit||cx>canvas.width) cx=-1;
                        else tw=Math.max(6,(e2[lane+1]-e2[lane])*0.75);
                    }
                }
                if(cx>=0){
                    const th=Math.max(3,4*scale);
                    ctx.save(); ctx.fillStyle='#00ff00';
                    ctx.fillRect(cx-tw/2,yD-th/2,tw,th);
                    ctx.strokeStyle='#004400'; ctx.lineWidth=Math.max(0.5,0.7*scale);
                    ctx.strokeRect(cx-tw/2,yD-th/2,tw,th); ctx.restore();
                }
            });
        }
    });
}

function loadImgFromB64(b64){
    try{
        var binary = atob(b64);
        var len = binary.length;
        var bytes = new Uint8Array(len);
        for(var i=0;i<len;i++) bytes[i] = binary.charCodeAt(i);
        var blob = new Blob([bytes], {type:'image/jpeg'});
        return URL.createObjectURL(blob);
    }catch(err){
        return "data:image/jpeg;base64,"+b64;
    }
}

function onRender(ev){
    const payload=(ev&&ev.detail)?ev.detail:ev;
    const args=payload&&payload.args?payload.args:null;
    if(!args) return;

    w1=args.w1||0; w2=args.w2||0; yGuia=args.y_guia||0;
    list1=args.list1||[]; list2=args.list2||[];
    skip1=args.skip1||0; skip2=args.skip2||0;

    if(!isReady){
        edges1=args.edges1||[];
        edges2=args.edges2||[];
        if(args.bandas_init) bandas=args.bandas_init;
    } else {
        if(args.force_edges){
            edges1=args.edges1||[];
            edges2=args.edges2||[];
        }
    }

    var newKey = (args.img_b64||"").substring(0,64) + "|" + (args.img_b64||"").length;
    if(window._lastImgKey !== newKey){
        window._lastImgKey = newKey;
        if(window._lastBlobUrl){
            try{ URL.revokeObjectURL(window._lastBlobUrl); }catch(e){}
        }
        var src = loadImgFromB64(args.img_b64||"");
        window._lastBlobUrl = src.indexOf("blob:")===0 ? src : null;
        
        img.onload = function(){
            originalW = img.width; originalH = img.height;
            scale = (args.largura_inicial||2600) / originalW;
            draw();
            if(S) S.setFrameHeight(700);
        };
        img.onerror = function(){
            toast("Falha ao carregar imagem do gel","error");
            if(S) S.setFrameHeight(700);
        };
        img.src = src;
    } else {
        draw();
        if(S) S.setFrameHeight(700);
    }
    isReady=true;
}

if(S){ 
    S.events.addEventListener("streamlit:render",onRender); 
    S.setComponentReady(); 
    S.setFrameHeight(700); 
}

container.addEventListener('wheel',(e)=>{
    e.preventDefault();
    const rect=canvas.getBoundingClientRect();
    const mx=e.clientX-rect.left, my=e.clientY-rect.top;
    const zf=e.deltaY<0?1.15:0.85;
    const ns=scale*zf;
    if(ns*originalW<400||ns*originalW>14000) return;
    scale=ns; draw();
    const cr=container.getBoundingClientRect();
    container.scrollLeft=mx*zf-(e.clientX-cr.left);
    container.scrollTop=my*zf-(e.clientY-cr.top);
},{passive:false});

document.addEventListener('keydown',(e)=>{
    if(e.code==='Space' && !e.repeat){
        spacePressed=true;
        if(!isEditingLanes || !draggingEdge) canvas.style.cursor='grab';
        e.preventDefault();
    }
    if(e.key==='Escape'){
        if(isEditingLanes){ toggleEditLanes(); return; }
        if(bandas.length>0&&confirm('Limpar todas as bandas?')){
            bandas=[]; draw(); sendData(); toast('Bandas removidas','info');
        }
    }
});
document.addEventListener('keyup',(e)=>{
    if(e.code==='Space'){
        spacePressed=false;
        if(!isPanning) canvas.style.cursor=isEditingLanes?(hoverEdge?'col-resize':'crosshair'):'crosshair';
    }
});

canvas.addEventListener('mousemove',(e)=>{
    const rect=canvas.getBoundingClientRect();
    const mX=(e.clientX-rect.left)/scale;

    if(isPanning){
        container.scrollLeft=scrollStartL-(e.clientX-panStartX);
        container.scrollTop=scrollStartT-(e.clientY-panStartY);
        return;
    }

    if(isEditingLanes){
        if(draggingEdge){
            if(draggingEdge.group===1) edges1[draggingEdge.idx]=mX;
            else edges2[draggingEdge.idx]=mX-w1;
            draw();
        } else {
            const oldH=hoverEdge?`${hoverEdge.group}-${hoverEdge.idx}`:null;
            hoverEdge=getNearestEdge(mX);
            const newH=hoverEdge?`${hoverEdge.group}-${hoverEdge.idx}`:null;
            if(!spacePressed) canvas.style.cursor=hoverEdge?'col-resize':'crosshair';
            if(oldH!==newH) draw();
        }
    }
});

canvas.addEventListener('mousedown',(e)=>{
    const rect=canvas.getBoundingClientRect();
    const mX=(e.clientX-rect.left)/scale;
    const mY=(e.clientY-rect.top)/scale;
    const totalInd=list1.length+list2.length;

    if(e.button===1 || (e.button===0 && spacePressed)){
        e.preventDefault();
        isPanning=true;
        panStartX=e.clientX; panStartY=e.clientY;
        scrollStartL=container.scrollLeft; scrollStartT=container.scrollTop;
        canvas.style.cursor='grabbing';
        return;
    }

    if(isEditingLanes){
        e.preventDefault(); e.stopPropagation();
        const nearest=getNearestEdge(mX);

        if(e.button===0){
            if(nearest){
                draggingEdge=nearest;
            } else {
                if(mX<=w1){
                    if(!edges1) edges1=[];
                    edges1.push(mX); edges1.sort((a,b)=>a-b);
                } else {
                    if(!edges2) edges2=[];
                    edges2.push(mX-w1); edges2.sort((a,b)=>a-b);
                }
                draw(); sendData();
            }
        } else if(e.button===2){
            if(nearest){
                if(nearest.group===1) edges1.splice(nearest.idx,1);
                else edges2.splice(nearest.idx,1);
                hoverEdge=null; draw(); sendData();
            }
        }
        return;
    }

    if(e.button===0){
        let clickedB=null;
        for(let b of bandas){ if(Math.abs(b.y-mY)<18/scale){clickedB=b;break;} }

        if(clickedB){
            let col=-1;
            if(mX<w1){
                const lane=laneFromEdges(edges1,mX);
                if(lane===-1){toast('Fora das colunas','warning');return;}
                const idx=lane-skip1;
                if(idx<0||idx>=list1.length){toast('Coluna ignorada (ladder/vazia)','info');return;}
                col=idx;
            } else {
                const lane=laneFromEdges(edges2,mX-w1);
                if(lane===-1){toast('Fora das colunas','warning');return;}
                const idx2=lane-skip2;
                if(idx2<0||idx2>=list2.length){toast('Coluna ignorada (ladder/vazia)','info');return;}
                col=list1.length+idx2;
            }
            if(col>=0&&col<totalInd){
                if(!clickedB.marks) clickedB.marks=new Array(totalInd).fill(0);
                if(clickedB.marks.length<totalInd){
                    clickedB.marks=clickedB.marks.concat(new Array(totalInd-clickedB.marks.length).fill(0));
                }
                clickedB.marks[col]=clickedB.marks[col]===1?0:1;
                const nome=col<list1.length?list1[col]:list2[col-list1.length];
                toast((clickedB.marks[col]?'Marcado: ':'Desmarcado: ')+nome, clickedB.marks[col]?'success':'warning');
                draw(); sendData();
            }
        } else {
            bandas.push({y:mY,marks:new Array(totalInd).fill(0)});
            bandas.sort((a,b)=>a.y-b.y);
            const li=bandas.findIndex(b=>b.y===mY);
            const letra=li<26?String.fromCharCode(97+li):`b${li}`;
            toast(`Linha "${letra}" criada!`,'info');
            draw(); sendData();
        }
    } else if(e.button===2){
        if(bandas.length===0) return;
        let ci=0,md=Math.abs(bandas[0].y-mY);
        for(let i=1;i<bandas.length;i++){const d=Math.abs(bandas[i].y-mY);if(d<md){md=d;ci=i;}}
        if(md<25/scale){
            const le=ci<26?String.fromCharCode(97+ci):`b${ci}`;
            bandas.splice(ci,1); toast(`Linha "${le}" excluida!`,'warning');
            draw(); sendData();
        }
    }
});

function endPanOrDrag(){
    if(isPanning){
        isPanning=false;
        canvas.style.cursor=spacePressed?'grab':(isEditingLanes?(hoverEdge?'col-resize':'crosshair'):'crosshair');
    }
    if(draggingEdge){
        if(edges1) edges1.sort((a,b)=>a-b);
        if(edges2) edges2.sort((a,b)=>a-b);
        draggingEdge=null;
        draw(); sendData();
    }
}
canvas.addEventListener('mouseup',endPanOrDrag);
canvas.addEventListener('mouseleave',endPanOrDrag);
canvas.addEventListener('contextmenu',e=>e.preventDefault());
canvas.addEventListener('auxclick',e=>{ if(e.button===1) e.preventDefault(); });
</script>
</body>
</html>
"""

COMP_VERSION = "v28_7_1_" + hashlib.md5(HTML_TEMPLATE.encode("utf-8")).hexdigest()[:10]

def _component_dir():
    candidates = [
        os.path.join(tempfile.gettempdir(), "ssr_canvas_component_" + COMP_VERSION),
        os.path.join(os.getcwd(), "ssr_canvas_component")
    ]
    last_err = None
    for comp_dir in candidates:
        try:
            os.makedirs(comp_dir, exist_ok=True)
            fp = os.path.join(comp_dir, "index.html")
            with open(fp, "w", encoding="utf-8") as f:
                f.write(HTML_TEMPLATE)
            if os.path.isfile(fp):
                return comp_dir
        except Exception as e:
            last_err = e
            continue
    raise RuntimeError(f"Nao foi possivel criar componente canvas: {last_err}")

@st.cache_resource
def get_interactive_canvas(version=COMP_VERSION):
    comp_dir = _component_dir()
    return components.declare_component("ssr_gel_canvas", path=comp_dir)


# ============================================================
#  ESTADO DA SESSAO
# ============================================================
defaults = {
    "offset_y2": 0,
    "rot_f1": 0.0,
    "rot_f2": 0.0,
    "escala_y2": 1.0,
    "pos_laser": 100,
    "auto_ok": False,
    "y1_det": None,
    "y2_det": None,
    "last_upload_id": None,
    "skip1": 1,
    "skip2": 0,
    "auto_update": True,
    "ordenar_ids_auto": True,
    "todas_matrizes": {},
    "ms_f1": [],
    "ms_f2": [],
    "brilho_val": 1.0,
    "contraste_val": 1.0,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ============================================================
#  CALLBACKS
# ============================================================

def _on_change_f1():
    if st.session_state.get("ordenar_ids_auto", True):
        st.session_state["ms_f1"] = ordenar_ids(st.session_state.get("ms_f1", []))

def _on_change_f2():
    if st.session_state.get("ordenar_ids_auto", True):
        st.session_state["ms_f2"] = ordenar_ids(st.session_state.get("ms_f2", []))

def _on_change_ordenar():
    if st.session_state.get("ordenar_ids_auto", True):
        st.session_state["ms_f1"] = ordenar_ids(st.session_state.get("ms_f1", []))
        st.session_state["ms_f2"] = ordenar_ids(st.session_state.get("ms_f2", []))


# ============================================================
#  INTERFACE DO APLICATIVO
# ============================================================
st.title("Sistema ISSR — Leitor de Gel Duplo v28.7.2")

nome_cultura_input = st.text_input(
    "Nome da Cultura / Material (ex: Milho, Feijao, Cafe):",
    value=st.session_state.get("nome_cultura_salva", "Minha_Cultura"),
    help="Este nome sera a pasta principal criada na sua Area de Trabalho (uso local)."
)
st.session_state["nome_cultura_salva"] = nome_cultura_input

atualizar_caminhos(nome_cultura_input)
garantir_pastas_base()

if not _is_cloud():
    st.caption(
        f"**Caminho exato no seu PC:** `{os.path.join(_get_desktop_path(), 'ssr_resultados', _nome_arquivo_seguro(nome_cultura_input))}`"
    )

aba1, aba2, aba3, aba4 = st.tabs(["1. Leitura do Gel", "2. Dendrograma UPGMA", "3. Exportar / Importar Excel", "4. Ajuda"])

# ─────────────────────────────────────────────────────────────
with aba1:
    st.header("Configuracao das Amostras")

    opcoes_ids = [str(i) for i in range(1, 301)] + ["L", "C", "C1", "C2", "M"]

    c1, c2 = st.columns([2, 5])
    with c1:
        nome_primer = st.text_input("Nome do Primer:", value="ISSR 19",
                                     help="Uma subpasta com este nome sera criada dentro da pasta da cultura.")
        st.session_state["auto_update"] = st.checkbox(
            "Atualizacao automatica",
            value=bool(st.session_state.get("auto_update", True)),
        )
        st.session_state["ordenar_ids_auto"] = st.checkbox(
            "Ordenar IDs no gel e dropdown",
            value=bool(st.session_state.get("ordenar_ids_auto", True)),
            on_change=_on_change_ordenar,
        )

    with c2:
        fc1, fc2 = st.columns(2)
        with fc1:
            st.multiselect(
                "Individuos Foto 1 (Esquerda):",
                options=opcoes_ids,
                key="ms_f1",
                on_change=_on_change_f1,
            )
        with fc2:
            st.multiselect(
                "Individuos Foto 2 (Direita):",
                options=opcoes_ids,
                key="ms_f2",
                on_change=_on_change_f2,
            )

    lista_f1 = list(st.session_state.get("ms_f1", []))
    lista_f2 = list(st.session_state.get("ms_f2", []))

    s1, s2, s3 = st.columns(3)
    with s1:
        if lista_f1:
            st.success(f"Foto 1: {len(lista_f1)} selecionados")
        else:
            st.warning("Foto 1: nenhum selecionado")
    with s2:
        if lista_f2:
            st.success(f"Foto 2: {len(lista_f2)} selecionados")
        else:
            st.warning("Foto 2: nenhum selecionado")
    with s3:
        lista_uni = list(dict.fromkeys(lista_f1 + lista_f2)) if (lista_f1 or lista_f2) else []
        st.info(f"Total unico de individuos: {len(lista_uni)}")

    st.divider()
    st.header("Upload das Fotos")
    cu1, cu2 = st.columns(2)
    with cu1:
        foto1 = st.file_uploader("FOTO 1 (Esquerda)", type=["png", "jpg", "jpeg", "tif"])
    with cu2:
        foto2 = st.file_uploader("FOTO 2 (Direita)", type=["png", "jpg", "jpeg", "tif"])

    if foto1 and foto2 and lista_f1 and lista_f2:
        img1, er1 = carregar_imagem(foto1)
        img2, er2 = carregar_imagem(foto2)
        if img1 is None:
            st.error(f"Foto 1: {er1}")
            st.stop()
        if img2 is None:
            st.error(f"Foto 2: {er2}")
            st.stop()
        upload_id = f"{foto1.name}_{foto1.size}_{foto2.name}_{foto2.size}"

        st.markdown("---")
        st.subheader("Calibracao Vertical")
        
        img1_work = img1.copy()
        img2_work = img2.copy()

        colA, colB = st.columns([2, 3])
        with colA:
            if st.button("ALINHAR AUTOMATICAMENTE", type="primary", use_container_width=True):
                off, ylaser, y1d, y2d = auto_calibrar(
                    img1_work, img2_work,
                    rot1=st.session_state["rot_f1"],
                    rot2=st.session_state["rot_f2"],
                    escala2=st.session_state["escala_y2"]
                )
                st.session_state.update({
                    "offset_y2": off, "pos_laser": ylaser,
                    "y1_det": y1d, "y2_det": y2d,
                    "auto_ok": True, "last_upload_id": upload_id
                })
                st.rerun()
            if st.session_state["last_upload_id"] != upload_id:
                off, ylaser, y1d, y2d = auto_calibrar(img1_work, img2_work)
                st.session_state.update({
                    "offset_y2": off, "pos_laser": ylaser,
                    "y1_det": y1d, "y2_det": y2d,
                    "auto_ok": True, "last_upload_id": upload_id
                })
                st.rerun()
        with colB:
            if st.session_state["auto_ok"]:
                st.success(
                    f"F1 Y={st.session_state['y1_det']} | "
                    f"F2 Y={st.session_state['y2_det']} | "
                    f"Offset={st.session_state['offset_y2']}px"
                )

        # -------------------------------------------------------------
        # Controles Manuais de Calibracao - Padronizado com +/-
        # -------------------------------------------------------------
        
        with st.expander("Controles Manuais de Calibracao"):
            cc1, cc2, cc3 = st.columns([2, 2, 3])

            with cc1:
                st.markdown("**Ajuste Fino Vertical**")
                b1, b2, b3, b4 = st.columns(4)
                if b1.button("+10", key="dy_p10"):
                    st.session_state["offset_y2"] -= 10; st.rerun()
                if b2.button("+1", key="dy_p1"):
                    st.session_state["offset_y2"] -= 1; st.rerun()
                if b3.button("-1", key="dy_m1"):
                    st.session_state["offset_y2"] += 1; st.rerun()
                if b4.button("-10", key="dy_m10"):
                    st.session_state["offset_y2"] += 10; st.rerun()
                st.session_state["offset_y2"] = st.number_input(
                    "Deslocamento Y:",
                    value=int(st.session_state["offset_y2"]),
                    step=1,
                    key="num_offset_y2"
                )

            with cc2:
                st.markdown("**Posicao do Laser**")
                lb1, lb2, lb3, lb4 = st.columns(4)
                if lb1.button("+10", key="laser_p10"):
                    st.session_state["pos_laser"] = min(500, int(st.session_state["pos_laser"]) + 10); st.rerun()
                if lb2.button("+1", key="laser_p1"):
                    st.session_state["pos_laser"] = min(500, int(st.session_state["pos_laser"]) + 1); st.rerun()
                if lb3.button("-1", key="laser_m1"):
                    st.session_state["pos_laser"] = max(5, int(st.session_state["pos_laser"]) - 1); st.rerun()
                if lb4.button("-10", key="laser_m10"):
                    st.session_state["pos_laser"] = max(5, int(st.session_state["pos_laser"]) - 10); st.rerun()
                st.session_state["pos_laser"] = st.number_input(
                    "Altura Linha Guia:",
                    min_value=5, max_value=500,
                    value=int(st.session_state["pos_laser"]),
                    step=1,
                    key="num_pos_laser"
                )

            with cc3:
                st.markdown("**Rotacao F1 (graus)**")
                r1a, r1b, r1c, r1d = st.columns(4)
                if r1a.button("+1.0", key="rf1_p10"):
                    st.session_state["rot_f1"] = round(min(10.0, float(st.session_state["rot_f1"]) + 1.0), 2); st.rerun()
                if r1b.button("+0.1", key="rf1_p1"):
                    st.session_state["rot_f1"] = round(min(10.0, float(st.session_state["rot_f1"]) + 0.1), 2); st.rerun()
                if r1c.button("-0.1", key="rf1_m1"):
                    st.session_state["rot_f1"] = round(max(-10.0, float(st.session_state["rot_f1"]) - 0.1), 2); st.rerun()
                if r1d.button("-1.0", key="rf1_m10"):
                    st.session_state["rot_f1"] = round(max(-10.0, float(st.session_state["rot_f1"]) - 1.0), 2); st.rerun()
                st.session_state["rot_f1"] = st.number_input(
                    "Rotacao F1:",
                    min_value=-10.0, max_value=10.0,
                    value=float(st.session_state["rot_f1"]),
                    step=0.1,
                    format="%.2f",
                    key="num_rot_f1"
                )

                st.markdown("**Rotacao F2 (graus)**")
                r2a, r2b, r2c, r2d = st.columns(4)
                if r2a.button("+1.0", key="rf2_p10"):
                    st.session_state["rot_f2"] = round(min(10.0, float(st.session_state["rot_f2"]) + 1.0), 2); st.rerun()
                if r2b.button("+0.1", key="rf2_p1"):
                    st.session_state["rot_f2"] = round(min(10.0, float(st.session_state["rot_f2"]) + 0.1), 2); st.rerun()
                if r2c.button("-0.1", key="rf2_m1"):
                    st.session_state["rot_f2"] = round(max(-10.0, float(st.session_state["rot_f2"]) - 0.1), 2); st.rerun()
                if r2d.button("-1.0", key="rf2_m10"):
                    st.session_state["rot_f2"] = round(max(-10.0, float(st.session_state["rot_f2"]) - 1.0), 2); st.rerun()
                st.session_state["rot_f2"] = st.number_input(
                    "Rotacao F2:",
                    min_value=-10.0, max_value=10.0,
                    value=float(st.session_state["rot_f2"]),
                    step=0.1,
                    format="%.2f",
                    key="num_rot_f2"
                )

                if st.button("Recalibrar", use_container_width=True, key="btn_recalibrar"):
                    off, ylaser, y1d, y2d = auto_calibrar(
                        img1_work, img2_work,
                        rot1=st.session_state["rot_f1"],
                        rot2=st.session_state["rot_f2"],
                        escala2=st.session_state["escala_y2"]
                    )
                    st.session_state.update({"offset_y2": off, "pos_laser": ylaser, "auto_ok": True})
                    st.rerun()

        # -------------------------------------------------------------
        # Ajustes Adicionais - Padronizado com +/-
        # -------------------------------------------------------------
        
        with st.expander("Ajustes Adicionais"):
            ce1, ce2, ce3 = st.columns(3)

            with ce1:
                st.markdown("**Brilho**")
                bb1, bb2, bb3, bb4 = st.columns(4)
                if bb1.button("+0.5", key="bri_p10"):
                    st.session_state["brilho_val"] = round(min(3.0, float(st.session_state["brilho_val"]) + 0.5), 2); st.rerun()
                if bb2.button("+0.1", key="bri_p1"):
                    st.session_state["brilho_val"] = round(min(3.0, float(st.session_state["brilho_val"]) + 0.1), 2); st.rerun()
                if bb3.button("-0.1", key="bri_m1"):
                    st.session_state["brilho_val"] = round(max(0.5, float(st.session_state["brilho_val"]) - 0.1), 2); st.rerun()
                if bb4.button("-0.5", key="bri_m10"):
                    st.session_state["brilho_val"] = round(max(0.5, float(st.session_state["brilho_val"]) - 0.5), 2); st.rerun()
                st.session_state["brilho_val"] = st.number_input(
                    "Brilho:",
                    min_value=0.5, max_value=3.0,
                    value=float(st.session_state["brilho_val"]),
                    step=0.1,
                    format="%.2f",
                    key="num_brilho"
                )
                brilho = st.session_state["brilho_val"]

            with ce2:
                st.markdown("**Contraste**")
                cb1, cb2, cb3, cb4 = st.columns(4)
                if cb1.button("+0.5", key="con_p10"):
                    st.session_state["contraste_val"] = round(min(3.0, float(st.session_state["contraste_val"]) + 0.5), 2); st.rerun()
                if cb2.button("+0.1", key="con_p1"):
                    st.session_state["contraste_val"] = round(min(3.0, float(st.session_state["contraste_val"]) + 0.1), 2); st.rerun()
                if cb3.button("-0.1", key="con_m1"):
                    st.session_state["contraste_val"] = round(max(0.5, float(st.session_state["contraste_val"]) - 0.1), 2); st.rerun()
                if cb4.button("-0.5", key="con_m10"):
                    st.session_state["contraste_val"] = round(max(0.5, float(st.session_state["contraste_val"]) - 0.5), 2); st.rerun()
                st.session_state["contraste_val"] = st.number_input(
                    "Contraste:",
                    min_value=0.5, max_value=3.0,
                    value=float(st.session_state["contraste_val"]),
                    step=0.1,
                    format="%.2f",
                    key="num_contraste"
                )
                contraste = st.session_state["contraste_val"]

            with ce3:
                escala_pct = int(round(float(st.session_state["escala_y2"]) * 100))

                st.markdown("**Escala F2 (%)**")
                eb1, eb2, eb3, eb4 = st.columns(4)
                if eb1.button("+5", key="esc_p10"):
                    escala_pct = min(120, escala_pct + 5)
                    st.session_state["escala_y2"] = escala_pct / 100.0; st.rerun()
                if eb2.button("+1", key="esc_p1"):
                    escala_pct = min(120, escala_pct + 1)
                    st.session_state["escala_y2"] = escala_pct / 100.0; st.rerun()
                if eb3.button("-1", key="esc_m1"):
                    escala_pct = max(80, escala_pct - 1)
                    st.session_state["escala_y2"] = escala_pct / 100.0; st.rerun()
                if eb4.button("-5", key="esc_m10"):
                    escala_pct = max(80, escala_pct - 5)
                    st.session_state["escala_y2"] = escala_pct / 100.0; st.rerun()
                escala_pct_new = st.number_input(
                    "Escala F2 (%):",
                    min_value=80, max_value=120,
                    value=int(escala_pct),
                    step=1,
                    key="num_escala_y2"
                )
                st.session_state["escala_y2"] = escala_pct_new / 100.0

                filtro_cor = st.selectbox(
                    "Filtro de Cor (Opcional):",
                    [
                        "Padrao (Cor Original)",
                        "Lilas/Roxo (Fundo Roxo / Bandas Brancas)",
                        "Preto e Branco (Fundo Preto)",
                        "Preto e Branco (Invertido - Fundo Branco)",
                        "Amarelo (Fundo Amarelo / Bandas Pretas)"
                    ],
                    index=0,
                    help="Escolha o visual do gel para facilitar a leitura."
                )

        with st.expander("Pular colunas iniciais (ladder/controle)"):
            col_s1, col_s2 = st.columns(2)
            with col_s1:
                st.session_state["skip1"] = st.number_input(
                    "F1 — colunas a pular", 0, 10, int(st.session_state["skip1"]), 1)
            with col_s2:
                st.session_state["skip2"] = st.number_input(
                    "F2 — colunas a pular", 0, 10, int(st.session_state["skip2"]), 1)

        # -------------------------------------------------------------
        # Aplicacao dos Filtros e Transformacoes
        # -------------------------------------------------------------
        if "Invertido" in filtro_cor:
            cor_fundo = (255, 255, 255)
        elif "Amarelo" in filtro_cor:
            cor_fundo = (0, 160, 255)
        elif "Lilas" in filtro_cor:
            cor_fundo = (150, 30, 90)
        else:
            cor_fundo = (0, 0, 0)
        
        img1_p = aplicar_filtro_bw(img1_work, filtro_cor, brilho, contraste)
        img2_p = aplicar_filtro_bw(img2_work, filtro_cor, brilho, contraste)
        
        img1_t = transformar_imagem(img1_p, st.session_state["rot_f1"], 1.0, 0, cor_fundo)
        img2_t = transformar_imagem(
            img2_p, st.session_state["rot_f2"],
            st.session_state["escala_y2"],
            st.session_state["offset_y2"], cor_fundo
        )
        h1, w1_img = img1_t.shape[:2]
        h2, w2_img = img2_t.shape[:2]
        if h2 != h1:
            nw2 = int(w2_img * (h1 / h2))
            img2_t = cv2.resize(img2_t, (nw2, h1), interpolation=cv2.INTER_AREA)
        cv2.line(img1_t, (w1_img - 2, 0), (w1_img - 2, h1), (255, 150, 0), 3)
        cv2.line(img2_t, (1, 0), (1, h1), (255, 150, 0), 3)
        img_unida = cv2.hconcat([img1_t, img2_t])
        
        _, buffer = cv2.imencode(".jpg", img_unida, [cv2.IMWRITE_JPEG_QUALITY, 96])
        img_b64 = base64.b64encode(buffer).decode()


        st.divider()
        st.subheader("Visualizador Interativo")

        largura_panoramica = st.slider("Tamanho Inicial (px):", 1000, 6000, 2600, 100)

        cols1 = len(lista_f1) + int(st.session_state["skip1"])
        cols2 = len(lista_f2) + int(st.session_state["skip2"])
        canvas_key = f"canvas_{upload_id}"

        if f"edges1_{canvas_key}" in st.session_state:
            edges1_init = st.session_state[f"edges1_{canvas_key}"]
        else:
            edges1_init = np.linspace(0, w1_img, cols1 + 1).round().astype(int).tolist()

        if f"edges2_{canvas_key}" in st.session_state:
            edges2_init = st.session_state[f"edges2_{canvas_key}"]
        else:
            edges2_init = np.linspace(0, img2_t.shape[1], cols2 + 1).round().astype(int).tolist()

        bandas_default = st.session_state.get(f"bandas_{canvas_key}", [])
        canvas_result = None

        try:
            interactive_canvas = get_interactive_canvas()
            canvas_result = interactive_canvas(
                img_b64=img_b64,
                w1=int(w1_img),
                w2=int(img2_t.shape[1]),
                y_guia=int(st.session_state["pos_laser"]),
                list1=lista_f1,
                list2=lista_f2,
                largura_inicial=int(largura_panoramica),
                skip1=int(st.session_state["skip1"]),
                skip2=int(st.session_state["skip2"]),
                edges1=edges1_init,
                edges2=edges2_init,
                bandas_init=bandas_default,
                default={"bandas": bandas_default, "calib_edges1": edges1_init, "calib_edges2": edges2_init},
                key=canvas_key
            )
        except Exception as e:
            st.error(f"Falha no canvas interativo: {e}")
            st.warning("Mostrando gel em modo fallback (sem marcacao por clique na imagem). Use a matriz abaixo.")
        
        if canvas_result is None:
            with st.expander("Preview estatico do gel (fallback visual)", expanded=False):
                st.image(cv2.cvtColor(img_unida, cv2.COLOR_BGR2RGB), use_container_width=True)

        auto_upd = bool(st.session_state.get("auto_update", True))
        if canvas_result and isinstance(canvas_result, dict):
            changed = False
            if "bandas" in canvas_result and canvas_result["bandas"] != st.session_state.get(f"bandas_{canvas_key}", []):
                st.session_state[f"bandas_{canvas_key}"] = canvas_result["bandas"]
                changed = True
            e1_res = canvas_result.get("calib_edges1")
            e2_res = canvas_result.get("calib_edges2")
            if e1_res is not None and e1_res != st.session_state.get(f"edges1_{canvas_key}"):
                st.session_state[f"edges1_{canvas_key}"] = e1_res
                changed = True
            if e2_res is not None and e2_res != st.session_state.get(f"edges2_{canvas_key}"):
                st.session_state[f"edges2_{canvas_key}"] = e2_res
                changed = True
            if changed and auto_upd:
                st.rerun()

        bandas_salvas = st.session_state.get(f"bandas_{canvas_key}", [])
        num_canvas_bands = len(bandas_salvas)

        st.divider()
        st.subheader("Matriz de Leitura")
        cm1, cm2 = st.columns([2, 4])
        with cm1:
            n_bandas = st.number_input(
                "Total de bandas:", min_value=1, max_value=100, value=max(1, num_canvas_bands))
        n_real = max(n_bandas, num_canvas_bands)
        with cm2:
            st.info(f"**{n_real} bandas** x **{len(lista_uni)} individuos**")

        letras_tabela = [chr(97 + i) if i < 26 else f"b{i}" for i in range(n_real)]
        key_mat = f"matriz_dados_{nome_primer}_{upload_id}"
        dados_bool = {ind: [False] * n_real for ind in lista_uni}
        for i, b in enumerate(bandas_salvas):
            if i >= n_real:
                break
            for col_idx, m in enumerate(b.get("marks", [])):
                if m == 1:
                    if col_idx < len(lista_f1):
                        ind_name = lista_f1[col_idx]
                    else:
                        idx2 = col_idx - len(lista_f1)
                        ind_name = lista_f2[idx2] if idx2 < len(lista_f2) else None
                    if ind_name and ind_name in dados_bool:
                        dados_bool[ind_name][i] = True

        df_at = pd.DataFrame(dados_bool, index=letras_tabela)
        df_at.index.name = "Banda"
        st.session_state[key_mat] = df_at
        df_ed = st.data_editor(
            st.session_state[key_mat],
            use_container_width=True, height=350,
            key=f"editor_{key_mat}"
        )

        st.session_state["matriz_final"] = df_ed
        st.session_state["primer_nome"] = nome_primer

        tm = df_ed.sum().sum()
        tc = n_real * len(lista_uni)
        taxa = (tm / tc * 100) if tc > 0 else 0
        cs1, cs2, cs3 = st.columns(3)
        cs1.metric("Total Marcado", f"{int(tm)}/{tc}")
        cs2.metric("Taxa de Presenca", f"{taxa:.1f}%")
        cs3.metric("Bandas x Individuos", f"{n_real} x {len(lista_uni)}")

        st.divider()
        st.subheader("Gerenciador de Primers (Multilocus)")

        col_m1, col_m2 = st.columns(2)
        with col_m1:
            st.markdown(
                f"**Salvar Primer Atual:**\nGrava a **matriz Excel** e o **gel processado**. "
            )
            if st.button(f"Salvar matriz e imagem do primer: {nome_primer}",
                         type="primary", use_container_width=True):
                atualizar_caminhos(st.session_state.get("nome_cultura_salva", nome_cultura_input))
                pasta_primer = garantir_pasta_primer(nome_primer)

                df_to_save = df_ed.copy()
                df_to_save.index = [f"{nome_primer}_{x}" for x in df_to_save.index]
                st.session_state["todas_matrizes"][nome_primer] = df_to_save

                ok_mat, path_mat = salvar_matriz_primer(nome_primer, df_to_save)

                e1_now = st.session_state.get(f"edges1_{canvas_key}", edges1_init)
                e2_now = st.session_state.get(f"edges2_{canvas_key}", edges2_init)
                ok_gel, info_gel = salvar_gel_conferencia(
                    img_bgr=img_unida, nome_primer=nome_primer,
                    bandas_salvas=bandas_salvas,
                    lista_f1=lista_f1, lista_f2=lista_f2,
                    edges1=e1_now, edges2=e2_now,
                    w1=int(w1_img),
                    skip1=int(st.session_state["skip1"]),
                    skip2=int(st.session_state["skip2"]),
                    y_guia=int(st.session_state["pos_laser"])
                )

                ok_bkp, path_bkp = salvar_backup_global()

                abrir_pasta(pasta_primer)

                linhas = [f"Primer **{nome_primer}** processado!",
                          f"Pasta: `{pasta_primer}`", ""]
                linhas.append(f"Matriz: `{path_mat}`" if ok_mat else f"Matriz (Erro): {path_mat}")
                linhas.append(f"Gel: `{info_gel}`" if ok_gel else f"Gel (Erro): {info_gel}")
                linhas.append(f"Backup: `{path_bkp}`" if ok_bkp else f"Backup (Erro): {path_bkp}")

                if ok_mat and ok_gel:
                    st.success("\n\n".join(linhas))
                elif ok_mat or ok_gel:
                    st.warning("\n\n".join(linhas))
                else:
                    st.error("\n\n".join(linhas))

            if st.button("Salvar apenas a foto do GEL com marcacoes", use_container_width=True):
                pasta_primer = garantir_pasta_primer(nome_primer)
                e1_now = st.session_state.get(f"edges1_{canvas_key}", edges1_init)
                e2_now = st.session_state.get(f"edges2_{canvas_key}", edges2_init)
                ok_gel, info_gel = salvar_gel_conferencia(
                    img_bgr=img_unida, nome_primer=nome_primer,
                    bandas_salvas=bandas_salvas,
                    lista_f1=lista_f1, lista_f2=lista_f2,
                    edges1=e1_now, edges2=e2_now,
                    w1=int(w1_img),
                    skip1=int(st.session_state["skip1"]),
                    skip2=int(st.session_state["skip2"]),
                    y_guia=int(st.session_state["pos_laser"])
                )
                if ok_gel:
                    abrir_pasta(pasta_primer)
                    st.success(f"Gel gravado em:\n`{info_gel}`")
                else:
                    st.error(f"Nao foi possivel salvar o gel: {info_gel}")

        with col_m2:
            st.markdown(
                "**Continuar a partir de Planilha Pronta:**\n"
                "Importe um Excel gerado anteriormente pelo App:"
            )
            arq_excel_a1 = st.file_uploader(
                "Selecionar planilha Excel para carregar dados:",
                type=["xlsx"], key="importar_excel_a1"
            )
            if arq_excel_a1 is not None:
                file_id = f"{arq_excel_a1.name}_{arq_excel_a1.size}"
                if st.session_state.get("last_import_a1") != file_id:
                    try:
                        dict_primers, erro = importar_excel_completo(arq_excel_a1.read())
                        if erro:
                            st.error(f"Falha ao ler arquivo: {erro}")
                        else:
                            st.session_state["todas_matrizes"].update(dict_primers)
                            salvar_backup_global()
                            st.session_state["last_import_a1"] = file_id
                            nomes_lidos = ", ".join(list(dict_primers.keys()))
                            st.session_state["msg_import_local_a1"] = (
                                f"{len(dict_primers)} primers importados:\n\n**{nomes_lidos}**"
                            )
                            st.toast(f"{len(dict_primers)} primers carregados!")
                    except Exception as ex:
                        st.error(f"Falha critica ao importar: {str(ex)}")
                if st.session_state.get("msg_import_local_a1"):
                    st.success(st.session_state["msg_import_local_a1"])

        st.divider()
        salvos = list(st.session_state["todas_matrizes"].keys())
        if salvos:
            st.caption(f"**{len(salvos)} Primers em memoria:** {', '.join(salvos)}")
            if st.button("Limpar toda a memoria de primers"):
                st.session_state["todas_matrizes"] = {}
                if os.path.exists(BACKUP_FILE):
                    try:
                        os.remove(BACKUP_FILE)
                    except:
                        pass
                st.rerun()
        else:
            st.caption("Nenhum primer na memoria temporaria.")

    elif not (lista_f1 and lista_f2):
        st.info("Selecione os individuos de cada foto para poder enviar as imagens.")
    else:
        st.info("Faca upload das duas fotos para iniciar.")


# ─────────────────────────────────────────────────────────────
with aba2:
    st.header("Analise de Diversidade Genetica (UPGMA Combinado)")

    salvos = list(st.session_state.get("todas_matrizes", {}).keys())

    if len(salvos) == 0:
        st.info("Salve ou Importe um primer na **Aba 1** para gerar o Dendrograma.")
    else:
        st.markdown("### Quais primers voce quer analisar juntos?")
        primers_selecionados = st.multiselect("Selecione os primers:", salvos, default=salvos)

        if primers_selecionados:
            dfs_to_combine = [st.session_state["todas_matrizes"][p] for p in primers_selecionados]
            df_combined = pd.concat(dfs_to_combine, axis=0).fillna(False)

            acessos = df_combined.columns.tolist()
            bandas_l = df_combined.index.tolist()
            mat_bin = df_combined.values.T.astype(int)

            st.info(
                f"Analisando **{len(bandas_l)} bandas** de **{len(primers_selecionados)} primer(s)** "
                f"para **{len(acessos)} individuos**."
            )

            if mat_bin.shape[0] >= 3 and np.sum(mat_bin) > 0:
                titulo_upgma = (
                    f"UPGMA — {len(primers_selecionados)} Primers Combinados "
                    f"({len(acessos)} Individuos)"
                )
                dj = calcular_jaccard(mat_bin)
                Z = fazer_upgma(dj)

                fig = plotar_dendrograma(Z, acessos, titulo_upgma)
                st.pyplot(fig)
                plt.close(fig)

                cd1, cd2 = st.columns(2)
                nome_arq = "Combinado" if len(primers_selecionados) > 1 else primers_selecionados[0]
                with cd1:
                    bf = io.BytesIO()
                    fd = plotar_dendrograma(Z, acessos, titulo_upgma)
                    fd.savefig(bf, format="png", dpi=300, bbox_inches="tight")
                    plt.close(fd)
                    st.download_button("PNG 300dpi", bf.getvalue(),
                                       f"dendrograma_{nome_arq}.png", "image/png",
                                       use_container_width=True)
                with cd2:
                    bp = io.BytesIO()
                    fp2 = plotar_dendrograma(Z, acessos, titulo_upgma)
                    fp2.savefig(bp, format="pdf", bbox_inches="tight")
                    plt.close(fp2)
                    st.download_button("PDF", bp.getvalue(),
                                       f"dendrograma_{nome_arq}.pdf", "application/pdf",
                                       use_container_width=True)

                st.divider()
                ce1, ce2 = st.columns(2)
                with ce1:
                    st.subheader("Genitores Contrastantes")
                    aux = dj.copy()
                    np.fill_diagonal(aux, -1)
                    ix = np.unravel_index(np.argmax(aux), aux.shape)
                    st.success(
                        f"**{acessos[ix[0]]}** x **{acessos[ix[1]]}**\n\n"
                        f"Dissimilaridade: **{aux[ix]:.3f}**"
                    )
                    with st.expander("Top 5"):
                        pares = sorted(
                            [(acessos[i], acessos[j], dj[i, j])
                             for i in range(len(acessos))
                             for j in range(i + 1, len(acessos))],
                            key=lambda x: x[2], reverse=True
                        )
                        for r, (a1, a2, d) in enumerate(pares[:5], 1):
                            st.write(f"**{r}.** {a1} x {a2} -> {d:.3f}")
                with ce2:
                    st.subheader("PIC por Banda")
                    ni = mat_bin.shape[0]
                    rows = []
                    for j in range(mat_bin.shape[1]):
                        p = np.sum(mat_bin[:, j]) / ni
                        pv = 1 - p ** 2 - (1 - p) ** 2
                        rows.append({
                            "Banda": bandas_l[j],
                            "Freq(1)": round(p, 3),
                            "PIC": round(pv, 3),
                            "Info": "OK" if pv > 0.25 else "F"
                        })
                    df_pic = pd.DataFrame(rows)
                    st.dataframe(
                        df_pic.style.background_gradient(subset=['PIC'], cmap='RdYlGn'),
                        use_container_width=True, hide_index=True
                    )
                    st.metric("PIC Medio Geral", f"{df_pic['PIC'].mean():.3f}")

                with st.expander("Matriz de Jaccard (Combinada)"):
                    df_j = pd.DataFrame(dj, index=acessos, columns=acessos)
                    st.dataframe(
                        df_j.style.background_gradient(cmap="Blues").format("{:.3f}"),
                        use_container_width=True
                    )
            else:
                st.warning("Marque pelo menos 3 individuos para gerar a arvore.")
        else:
            st.warning("Selecione pelo menos um primer para analise.")


# ─────────────────────────────────────────────────────────────
with aba3:
    st.header("Exportar / Importar Matrizes SSR")

    with st.expander("IMPORTAR DE PLANILHA SSR EXISTENTE", expanded=True):
        st.markdown("Faca o upload de uma planilha gerada por esta ferramenta para restaurar e continuar:")
        arq_excel_a3 = st.file_uploader(
            "Fazer upload do arquivo SSR_Combinado.xlsx:",
            type=["xlsx"], key="importar_excel_a3"
        )
        if arq_excel_a3 is not None:
            file_id = f"{arq_excel_a3.name}_{arq_excel_a3.size}"
            if st.session_state.get("last_import_a3") != file_id:
                try:
                    dict_primers, erro = importar_excel_completo(arq_excel_a3.read())
                    if erro:
                        st.error(f"Nao foi possivel carregar: {erro}")
                    else:
                        st.session_state["todas_matrizes"].update(dict_primers)
                        salvar_backup_global()
                        st.session_state["last_import_a3"] = file_id
                        nomes_lidos = ", ".join(list(dict_primers.keys()))
                        st.session_state["msg_import_local_a3"] = (
                            f"Planilha restaurada! {len(dict_primers)} primers:\n\n**{nomes_lidos}**"
                        )
                        st.toast(f"{len(dict_primers)} primers carregados!")
                except Exception as ex:
                    st.error(f"Erro inesperado: {str(ex)}")
            if st.session_state.get("msg_import_local_a3"):
                st.success(st.session_state["msg_import_local_a3"])

    st.divider()
    st.subheader("Exportar Planilha Excel Atual")
    salvos = list(st.session_state.get("todas_matrizes", {}).keys())

    if len(salvos) == 0:
        st.info("Salve pelo menos um primer ou importe uma planilha para exportar.")
    else:
        st.markdown("### Selecione os primers para juntar no Excel:")
        primers_export = st.multiselect(
            "Exportar os seguintes primers:", salvos, default=salvos, key="export_multiselect"
        )
        if primers_export:
            dict_export = {p: st.session_state["todas_matrizes"][p] for p in primers_export}
            dfs = [st.session_state["todas_matrizes"][p] for p in primers_export]
            df_combined = pd.concat(dfs, axis=0).fillna(False)
            acessos = df_combined.columns.tolist()
            mat_bin = df_combined.values.T.astype(int)
            dj = None
            if mat_bin.shape[0] >= 3 and np.sum(mat_bin) > 0:
                dj = calcular_jaccard(mat_bin)
            nome_export = "Combinado" if len(primers_export) > 1 else primers_export[0]
            st.write("Matriz Binaria em blocos · Jaccard · Arial 11 Centrado · TXT Transposto")
            st.divider()

            buf = exportar_excel_completo(dict_export, nome_export, dj, acessos)
            txt_buf = gerar_txt_transposto(dict_export)
            nome_cult = st.session_state.get("nome_cultura_salva", "Minha_Cultura")
            nome_cult_seguro = _nome_arquivo_seguro(nome_cult)

            cb1, cb2, cb3 = st.columns(3)
            with cb1:
                st.download_button(
                    "Baixar Excel",
                    buf,
                    f"SSR_{nome_export}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with cb2:
                st.download_button(
                    f"Baixar TXT (Ent-{nome_cult_seguro}.txt)",
                    txt_buf.encode("ascii", errors="strict"),
                    f"Ent-{nome_cult_seguro}.txt",
                    "text/plain",
                    use_container_width=True,
                    help="Matriz transposta: so 0 e 1, um espaco entre valores, formato GENES.",
                )
            with cb3:
                if st.button("Salvar em 'Codificacao de bandas'",
                             type="primary", use_container_width=True):
                    ok, info, path_xlsx, path_txt = salvar_codificacao_bandas(
                        dict_export, nome_cult, dj, acessos
                    )
                    if ok:
                        abrir_pasta(info)
                        st.success(
                            f"Arquivos salvos com sucesso!\n\n"
                            f"Pasta: `{info}`\n\n"
                            f"Excel: `{path_xlsx}`\n\n"
                            f"TXT: `{path_txt}`"
                        )
                    else:
                        st.error(f"Falha ao salvar: {info}")

            st.success(
                f"Pronto: **{len(df_combined)} bandas** x **{len(acessos)} individuos** "
                f"· TXT no formato `Ent-{nome_cult_seguro}.txt`"
            )
            with st.expander("Previa do TXT transposto (primeiras linhas)"):
                preview_lines = [ln for ln in txt_buf.replace("\r\n", "\n").strip().split("\n") if ln != ""]
                n_show = min(8, len(preview_lines))
                st.code("\n".join(preview_lines[:n_show]), language="text")
                st.caption(
                    f"Total: **{len(preview_lines)} linhas** (individuos) x "
                    f"**{len(preview_lines[0].split()) if preview_lines else 0} colunas** (bandas). "
                    f"Apenas 0 e 1, separados por um espaco (formato GENES)."
                )


# ─────────────────────────────────────────────────────────────
with aba4:
    st.header("Guia de Uso")
    st.markdown(
        "### Estrutura de Pastas Criada Automaticamente\n\n"
        "Ao salvar um primer localmente, o sistema gera esta hierarquia na sua **Area de Trabalho**:"
    )

    st.code(
        "Desktop/\n"
        "└── ssr_resultados/                       ← Pasta principal (fixa)\n"
        "    └── Nome_da_Cultura/                  ← Nome digitado no topo do App\n"
        "        ├── backup_<Cultura>.pkl          ← Backup global de toda a cultura\n"
        "        ├── Codificacao de bandas/        ← Pasta unificada de exportacao\n"
        "        │   ├── SSR_Combinado.xlsx        ← Excel com todos os primers\n"
        "        │   └── Ent-<Cultura>.txt         ← TXT transposto (so 0 e 1)\n"
        "        ├── ISSR_19/                      ← Pasta do primer ISSR 19\n"
        "        │   ├── Matriz_ISSR_19.xlsx\n"
        "        │   └── gel_ISSR_19_<data>.jpg\n"
        "        ├── ISSR_20/                      ← Pasta do primer ISSR 20\n"
        "        │   ├── Matriz_ISSR_20.xlsx\n"
        "        │   └── gel_ISSR_20_<data>.jpg\n"
        "        └── ...",
        language="text",
    )

    st.markdown(
        "### Controles Rapidos do Visualizador\n\n"
        "| Acao | Como fazer |\n"
        "|------|------------|\n"
        "| **Zoom** | Roda do mouse |\n"
        "| **Mover a Imagem** | Botao do meio ou Espaco + arrastar |\n"
        "| **Criar Nova Coluna** | Ativar modo azul e clicar com Botao Esquerdo |\n"
        "| **Mover Coluna** | Segurar e arrastar a linha |\n"
        "| **Apagar Linha de Coluna** | Botao Direito sobre a linha |\n"
        "| **Criar Banda** | Botao Esquerdo no gel (fora das bandas) |\n"
        "| **Marcar Banda (0/1)** | Botao Esquerdo sobre o marcador verde |\n"
        "| **Deletar Banda** | Botao Direito sobre a banda |\n\n"
        "### Filtros de Cor Disponiveis\n"
        "- **Padrao (Cor Original)** — igual a foto enviada\n"
        "- **Lilas/Roxo** — fundo roxo com bandas brancas (estilo UV)\n"
        "- **P&B Fundo Preto** — escala de cinza padrao\n"
        "- **P&B Invertido** — negativo, fundo branco\n"
        "- **Amarelo (Fundo Ouro / Bandas Pretas)** — estilo nitrato de prata\n\n"
        "### Arquivo TXT Transposto (Ent-<Cultura>.txt)\n"
        "- Gerado automaticamente na Aba 3\n"
        "- Contem apenas valores `0` e `1` separados por um espaco\n"
        "- Sem nomes de primer, banda ou individuo\n"
        "- Cada linha representa 1 individuo\n"
        "- Cada coluna representa 1 banda\n"
        "- Formato ASCII + CRLF, compativel com GENES\n"
        "- Ideal para importacao em softwares de analise (GENES, R, PAST, etc.)\n\n"
        "### Ordenacao de IDs\n"
        "- Ao adicionar um item -> e inserido no lugar correto (ordenado)\n"
        "- Ao remover um item -> mantem a ordem\n"
        "- Ao ativar/desativar o checkbox -> dropdown reordena instantaneamente\n\n"
        "### Como salvar corretamente na Nuvem\n"
        "1. Digite o **Nome da Cultura** no topo (ex: Milho, Feijao).\n"
        "2. Digite o **Nome do Primer** (ex: ISSR 19).\n"
        "3. Faca a marcacao no gel.\n"
        "4. Clique em **Salvar matriz e imagem do primer** (para fixar na memoria).\n"
        "5. Va ate a **Aba 3**, selecione o primer e baixe o Excel ou o arquivo TXT nos botoes indicados.\n"
    )
    st.divider()
    st.success("ISSR Pro v28.7.2 — Login Clean Branco e Otimizado!")